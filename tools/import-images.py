#!/usr/bin/env python3
"""商品の実写画像を assets/product-images.json に取り込む。

型番から画像を「検索して拾ってくる」ことはしない。ECモールや検索エンジンの画像は
他社が撮影したもので、自社サイトへの転載は著作権侵害になるため。
使えるのは、メーカーまたは仕入先から販売目的での利用を許諾された画像だけ。

想定している入手経路は2つ:

  1. 仕入先の商品マスタ（型番列と画像URL列を持つCSV / Excel）
     SB C&S・ダイワボウ情報システムなどが販売店向けに配布しているもの

         python3 tools/import-images.py --from-file 商品マスタ.xlsx \\
             --model-col 型番 --url-col 画像URL --credit エレコム

  2. メーカーの販売店向け画像ダウンロード（型番名の画像ファイルが入ったフォルダ）

         python3 tools/import-images.py --from-dir ~/Downloads/elecom-images --credit エレコム

  3. 手元でコピーしたURLを貼り付ける（自社ショップの管理画面、メーカーの製品ページなど）

         python3 tools/import-images.py --paste --credit 自社撮影
         （標準入力に「型番<タブ/カンマ/空白>URL」を1行ずつ）

  4. 自社ショップ（Supabase の shop_products）から型番一致で取り込む

         python3 tools/import-images.py --from-shop

     ただしショップは中古品を1点ずつ扱っており、写真はその個体そのもの。
     新品として掲載する在庫に流用すると実物と食い違うため、
     一致した型番を一覧表示するだけで止まる。取り込むには --yes が要る。

--download を付けると、URLの画像を assets/products/ に保存してそちらを参照する。
直リンクは相手側の帯域を使い、URL変更で静かに壊れるので、本番はこちらを推奨。
（このスクリプトを外部へ接続できる環境で実行する必要がある）

    python3 tools/import-images.py --status   # 型番ごとの画像の有無を集計
"""

import argparse
import json
import re
import shutil
import sys
import urllib.request
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
IMAGES_JSON = ROOT / "assets" / "product-images.json"
INVENTORY = ROOT / "assets" / "inventory-data.js"
STORE = ROOT / "assets" / "products"

EXT_OK = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}


def load_images():
    return json.loads(IMAGES_JSON.read_text(encoding="utf-8"))


def save_images(data):
    data["updated"] = date.today().isoformat()
    data["images"] = dict(sorted(data["images"].items()))
    IMAGES_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_inventory():
    """在庫データの型番一覧。取り込んだ画像が在庫の型番と噛み合っているか照合する。"""
    src = INVENTORY.read_text(encoding="utf-8")
    data = json.loads(src[src.index("{"):].rstrip().rstrip(";"))
    return data["records"]


def rows_of(path):
    """CSV / Excel を「見出し行 + データ行」として読む。"""
    if path.suffix.lower() in {".csv", ".tsv", ".txt"}:
        import csv
        sep = "\t" if path.suffix.lower() == ".tsv" else ","
        # 仕入先のCSVはShift_JISのことが多い
        for encoding in ("utf-8-sig", "cp932"):
            try:
                with path.open(encoding=encoding, newline="") as fh:
                    return [r for r in csv.reader(fh, delimiter=sep)]
            except UnicodeDecodeError:
                continue
        sys.exit(f"{path.name} の文字コードを判別できませんでした")

    import openpyxl
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = [[("" if c is None else str(c)).strip() for c in r]
            for r in book[book.sheetnames[0]].iter_rows(values_only=True)]
    book.close()
    return rows


def find_header(rows, *names):
    """見出し行と、指定した列名の位置を探す。列の並びは配布元ごとに違うため。"""
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip() for c in row]
        found = {}
        for name in names:
            for j, cell in enumerate(cells):
                if name and name in cell:
                    found[name] = j
                    break
        if len(found) == len(names):
            return i, found
    return None, {}


def add(data, model, src, credit, note=""):
    model = model.strip()
    if not model:
        return False
    entry = {"src": src, "credit": credit}
    if note:
        entry["note"] = note
    data["images"][model] = entry
    return True


def download(url, model):
    """画像をローカルに保存して、直リンクを避ける。"""
    STORE.mkdir(parents=True, exist_ok=True)
    ext = Path(url.split("?")[0]).suffix.lower()
    if ext not in EXT_OK:
        ext = ".jpg"
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", model)
    dest = STORE / f"{safe}{ext}"
    req = urllib.request.Request(url, headers={"User-Agent": "8ec-image-import/1.0"})
    with urllib.request.urlopen(req, timeout=30) as res, dest.open("wb") as fh:
        shutil.copyfileobj(res, fh)
    return f"assets/products/{dest.name}"


def from_file(args, data):
    path = Path(args.from_file)
    rows = rows_of(path)
    head, cols = find_header(rows, args.model_col, args.url_col)
    if head is None:
        sys.exit(f"見出し「{args.model_col}」「{args.url_col}」が見つかりません。"
                 f"先頭行: {rows[0][:8] if rows else '(空)'}")

    added = failed = 0
    for row in rows[head + 1:]:
        if len(row) <= max(cols.values()):
            continue
        model = str(row[cols[args.model_col]]).strip()
        url = str(row[cols[args.url_col]]).strip()
        if not model or not url.startswith("http"):
            continue
        src = url
        if args.download:
            try:
                src = download(url, model)
            except Exception as exc:  # 1件の失敗で全体を止めない
                print(f"  取得失敗 {model}: {exc}")
                failed += 1
                continue
        added += add(data, model, src, args.credit)
    print(f"{path.name}: {added}件を登録" + (f"（{failed}件は取得失敗）" if failed else ""))


def from_dir(args, data):
    src_dir = Path(args.from_dir).expanduser()
    if not src_dir.is_dir():
        sys.exit(f"{src_dir} はフォルダではありません")
    STORE.mkdir(parents=True, exist_ok=True)

    added = 0
    for path in sorted(src_dir.iterdir()):
        if path.suffix.lower() not in EXT_OK:
            continue
        model = path.stem.strip()          # ファイル名がそのまま型番である前提
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", model)
        dest = STORE / f"{safe}{path.suffix.lower()}"
        shutil.copyfile(path, dest)
        added += add(data, model, f"assets/products/{dest.name}", args.credit)
    print(f"{src_dir.name}: {added}件を assets/products/ に取り込みました")


def from_paste(args, data):
    """「型番 URL」を1行ずつ貼り付けて登録する。

    ショップの管理画面やメーカーの製品ページからコピーしたURLを、
    そのまま流し込むための入口。区切りはタブ・カンマ・空白のいずれでもよい。
    """
    if sys.stdin.isatty():
        print("型番とURLを「型番<タブ/カンマ/空白>URL」の形式で貼り付けてください。")
        print("入力し終えたら Ctrl-D（Windows は Ctrl-Z → Enter）。\n")

    added = skipped = failed = 0
    for line in sys.stdin:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"[\t,]|\s{2,}|\s+(?=https?://)", line, maxsplit=1)
        if len(parts) < 2:
            print(f"  読み取れません: {line[:60]}")
            skipped += 1
            continue
        model, url = parts[0].strip(), parts[1].strip()
        if not url.startswith("http") and not url.startswith("assets/"):
            print(f"  URLではありません: {model} → {url[:40]}")
            skipped += 1
            continue
        src = url
        if args.download and url.startswith("http"):
            try:
                src = download(url, model)
            except Exception as exc:
                print(f"  取得失敗 {model}: {exc}")
                failed += 1
                continue
        if add(data, model, src, args.credit):
            added += 1

    print(f"{added}件を登録"
          + (f"（{skipped}件は形式不明）" if skipped else "")
          + (f"（{failed}件は取得失敗）" if failed else ""))


def from_shop(args, data):
    """自社ショップ（Supabase の公開ビュー shop_products）から型番一致で取り込む。

    ショップは中古品を1点ずつ扱っているため、写真はその個体そのもの。
    新品として売る在庫の商品画像に流用すると実物と食い違うので、
    型番が一致したものを一覧で出し、取り込むかどうかは人が判断する。
    """
    url = (args.shop_url.rstrip("/")
           + "/rest/v1/shop_products?select=model,image_url,images&limit=1000")
    req = urllib.request.Request(url, headers={
        "apikey": args.shop_key,
        "Authorization": f"Bearer {args.shop_key}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as res:
            rows = json.load(res)
    except Exception as exc:
        sys.exit(f"ショップの商品を取得できませんでした: {exc}\n"
                 f"（{args.shop_url} へ接続できる環境で実行してください）")

    records = load_inventory()
    inventory_models = {r["model"].upper(): r for r in records if r.get("model")}

    matched, no_photo = [], 0
    for row in rows:
        model = (row.get("model") or "").strip()
        shots = row.get("images") or []
        src = shots[0] if isinstance(shots, list) and shots else row.get("image_url")
        if not model:
            continue
        if not src:
            no_photo += 1
            continue
        hit = inventory_models.get(model.upper())
        if hit:
            matched.append((hit["model"], src, hit.get("maker") or "", hit.get("name") or ""))

    print(f"ショップの商品 {len(rows)}件 中、写真あり {len(rows) - no_photo}件。")
    print(f"法人在庫の型番と一致したもの: {len(matched)}件\n")
    if not matched:
        print("一致する型番がありませんでした。ショップは中古品、法人在庫は新品の"
              "ディストリビューター商材なので、型番が重なることは多くありません。")
        return

    for model, src, maker, name in matched[:30]:
        print(f"  {model:<20} {maker:<12} {name[:28]}")
    if len(matched) > 30:
        print(f"  ... 他 {len(matched) - 30}件")

    print("\nショップの写真はその中古個体を撮ったものです。新品として掲載する在庫に"
          "そのまま使うと実物と食い違います。")
    if not args.yes:
        print("取り込む場合は --yes を付けて再実行してください。")
        return

    for model, src, _, _ in matched:
        final = src
        if args.download and src.startswith("http"):
            try:
                final = download(src, model)
            except Exception as exc:
                print(f"  取得失敗 {model}: {exc}")
                continue
        add(data, model, final, args.credit, note="自社ショップの中古個体を撮影したもの")
    print(f"\n{len(matched)}件を登録しました。")


def status(data):
    records = load_inventory()
    images = data["images"]
    by_maker = {}
    for rec in records:
        maker = rec.get("maker") or "(メーカー未設定)"
        has = bool(rec.get("image")) or rec["model"] in images
        got, total = by_maker.get(maker, (0, 0))
        by_maker[maker] = (got + (1 if has else 0), total + 1)

    covered = sum(g for g, _ in by_maker.values())
    print(f"実写あり {covered} / {len(records)} 件（残りはイメージ図で表示）\n")
    print(f"{'メーカー':<24}{'実写':>6}{'全体':>7}")
    for maker, (got, total) in sorted(by_maker.items(), key=lambda kv: -kv[1][1]):
        print(f"{maker:<24}{got:>6}{total:>7}")

    unknown = [m for m in images if not any(r["model"] == m for r in records)]
    if unknown:
        print(f"\n在庫に無い型番の画像が {len(unknown)} 件あります: {', '.join(unknown[:5])}"
              + (" ..." if len(unknown) > 5 else ""))


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--from-file", help="型番列と画像URL列を持つ商品マスタ（CSV / Excel）")
    ap.add_argument("--from-dir", help="ファイル名が型番になっている画像フォルダ")
    ap.add_argument("--paste", action="store_true",
                    help="「型番 URL」を1行ずつ貼り付けて登録する（標準入力）")
    ap.add_argument("--from-shop", action="store_true",
                    help="自社ショップ（Supabase の shop_products）から型番一致で取り込む")
    ap.add_argument("--model-col", default="型番", help="型番の列見出し（--from-file 用）")
    ap.add_argument("--url-col", default="画像URL", help="画像URLの列見出し（--from-file 用）")
    ap.add_argument("--credit", default="メーカー提供", help="画像に添えるクレジット表記")
    ap.add_argument("--download", action="store_true", help="画像をローカルに保存して直リンクを避ける")
    ap.add_argument("--status", action="store_true", help="型番ごとの画像の有無を集計する")
    ap.add_argument("--yes", action="store_true", help="確認せずに取り込む（--from-shop 用）")
    ap.add_argument("--shop-url", default="https://htglvascsuqkixpmclwr.supabase.co",
                    help="ショップのSupabase URL")
    ap.add_argument("--shop-key", default="sb_publishable_yZCcrwdqjuf0u_5WBWlHIw_AxdvteEV",
                    help="ショップの公開APIキー（anon）")
    args = ap.parse_args()

    data = load_images()

    if args.status:
        status(data)
        return
    if args.from_file:
        from_file(args, data)
    elif args.from_dir:
        from_dir(args, data)
    elif args.paste:
        if args.credit == "メーカー提供":
            args.credit = "自社撮影"
        from_paste(args, data)
    elif args.from_shop:
        if args.credit == "メーカー提供":
            args.credit = "自社撮影"
        from_shop(args, data)
    else:
        ap.print_help()
        return

    save_images(data)
    print(f"assets/product-images.json を更新しました（計 {len(data['images'])} 件）")


if __name__ == "__main__":
    # head などに繋いだときに落ちないようにする
    try:
        import signal
        signal.signal(signal.SIGPIPE, signal.SIG_DFL)
    except (ImportError, AttributeError, ValueError):
        pass
    main()
